import os
import django
import click
import MySQLdb

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()

from onlineapp.models import *

@click.group()
@click.option('-u', '--user', type=str, default='root', help='username')
@click.option('-p', '--pwd', type=str, default='8897', help='password')
@click.pass_context
def mysql_wb(ctx, user, pwd):
    ctx.obj['USER'] = user
    ctx.obj['PWD'] = pwd

@click.command(help='creates a database with the given name and creates the tables- students and marks')
@click.argument('dbname', type=str, required=True)
@click.pass_context
def createdb(ctx, dbname):
    conn = MySQLdb.connect('localhost', ctx.obj['USER'], ctx.obj['PWD'])
    cur = conn.cursor()
    cur.execute('drop database if exists %s' % dbname)
    cur.execute('create database %s' % dbname)
    conn = MySQLdb.connect('localhost', ctx.obj['USER'], ctx.obj['PWD'], dbname)
    cur = conn.cursor()
    sql_create_student = '''
    create table students(
        name varchar(50),
        college varchar(10),
        email varchar(50),
        dbname varchar(20),
        primary key(college, dbname))
        '''
    sql_create_marks = '''
    create table marks(
        college varchar(10),
        dbname varchar(20),
        transform int,
        from_custom_base26 int,
        get_pig_latin int,
        top_chars int,
        Total int,
        foreign key(college, dbname) references students(college, dbname))
        '''
    cur.execute(sql_create_student)
    cur.execute(sql_create_marks)
    print('Created database %s successfully' % dbname)

@click.command(help='drops the database with the given name')
@click.argument('dbname', type=str, required=True)
@click.pass_context
def dropdb(ctx, dbname):
    conn = MySQLdb.connect('localhost', ctx.obj['USER'], ctx.obj['PWD'])
    cur = conn.cursor()
    cur.execute('drop database if exists %s' % dbname)
    print('Dropped database %s successfully' % dbname)

@click.command(help='imports data from excel file to the database tables')
@click.argument('dbname', type=str, required=True)
@click.argument('students_excel', type=str, default='students.xlsx')
@click.argument('marks_excel', type=str, default='marks.xlsx')
@click.pass_context
def importdata(ctx, dbname, students_excel, marks_excel):
    from openpyxl import load_workbook

    conn = MySQLdb.connect('localhost', ctx.obj['USER'], ctx.obj['PWD'], dbname)
    cur = conn.cursor()

    wb = load_workbook(students_excel)

    # importing college data first as student depends on college
    ws = wb['Colleges']
    for college_name, acronym, location, contact in ws.iter_rows(min_row=2):
        if not None in [college_name.value, acronym.value, location.value, contact.value]:
            college = College()
            college.name = college_name.value.strip()
            college.acronym = acronym.value.strip()
            college.location = location.value.strip()
            college.contact = contact.value.strip()
            college.save()

    # importing student data who are active
    ws = wb['Current']
    for name, college, email, dbnames in ws.iter_rows(min_row=2):
        if not None in [name.value, college.value, email.value, dbnames.value]:
            student = Student(dropped_out = False)
            student.name = name.value.strip()
            student.college = College.objects.get(acronym=college.value.strip())
            student.email = email.value.strip()
            student.db_folder = dbnames.value.strip().lower()
            student.save()

    # importing student data who have dropped out
    ws = wb['Deletions']
    for name, college, email, dbnames in ws.iter_rows(min_row=2):
        if not None in [name.value, college.value, email.value, dbnames.value]:
            student = Student(dropped_out = True)
            student.name = name.value.strip()
            student.college = College.objects.get(acronym=college.value.strip())
            student.email = email.value.strip()
            student.db_folder = dbnames.value.strip().lower()
            student.save()

    # importing mocktest1 marks
    wb = load_workbook(marks_excel)
    ws = wb.active
    for name, p1, p2, p3, p4, total in ws.iter_rows(min_row=2):
        if not None in [name.value, p1.value, p2.value, p3.value, p4.value, total.value]:
            marks = MockTest1()
            marks.student = Student.objects.get(db_folder=name.value.split("_")[2])
            marks.problem1 = p1.value
            marks.problem2 = p2.value
            marks.problem3 = p3.value
            marks.problem4 = p4.value
            marks.total = total.value
            marks.save()

    print('Imported data to database %s successfully' % dbname)

@click.command(help='prints the stats of all the colleges')
@click.argument('dbname', type=str, required=True)
@click.pass_context
def collegestats(ctx, dbname):
    conn = MySQLdb.connect('localhost', ctx.obj['USER'], ctx.obj['PWD'], dbname)
    cur = conn.cursor()
    sql_console_ouput = '''
    select s.college, count(*) as count, min(Total) as min, avg(Total) as avg, max(Total) as max
    from students s
    inner join marks m
    on s.college = m.college and s.dbname = m.dbname
    group by s.college
    '''
    cur.execute(sql_console_ouput)
    print('{:>10s}  {:>7s}  {:>7s}  {:>7s}  {:>7s}'.format('college', 'count', 'min', 'avg', 'max'))
    out_str = '{:>10s}  {:7d}  {:7d}  {:7d}  {:7d}'
    for college, count, min, avg, max in cur.fetchall():
        print(out_str.format(college, count, min, int(avg), max))

if __name__ == '__main__':
    mysql_wb.add_command(createdb)
    mysql_wb.add_command(dropdb)
    mysql_wb.add_command(importdata)
    mysql_wb.add_command(collegestats)
    mysql_wb(obj=dict())